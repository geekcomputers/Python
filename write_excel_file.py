import xlwt # type: ignore
import openpyxl # type: ignore

# Workbook is created
xlwt_wb = xlwt.Workbook()

"""
we can also use of json object format for this file or code,
for the index we can use (for loop) and for data use json object.
example of json object:
{
    "data":[
        "ISBT DEHRADUN".
        "SHASTRADHARA",
        "CLEMEN TOWN",
        "RAJPUR ROAD",
        "CLOCK TOWER",
        "ISBT DEHRADUN",
        "SHASTRADHARA",
        "CLEMEN TOWN",
        "RAJPUR ROAD",
        "CLOCK TOWER"
    ]
}
"""

# add_sheet is used to create sheet.
sheet1 = xlwt_wb.add_sheet("Sheet 1")

sheet1.write(1, 0, "ISBT DEHRADUN")
sheet1.write(2, 0, "SHASTRADHARA")
sheet1.write(3, 0, "CLEMEN TOWN")
sheet1.write(4, 0, "RAJPUR ROAD")
sheet1.write(5, 0, "CLOCK TOWER")
sheet1.write(0, 1, "ISBT DEHRADUN")
sheet1.write(0, 2, "SHASTRADHARA")
sheet1.write(0, 3, "CLEMEN TOWN")
sheet1.write(0, 4, "RAJPUR ROAD")
sheet1.write(0, 5, "CLOCK TOWER")

xlwt_wb.save("xlwt example.xls")

# Workbook is created
openpyxl_wb = openpyxl.Workbook()

# create_sheet is used to create sheet.
sheet1 = openpyxl_wb.create_sheet("Sheet 1", index=0)

sheet1.cell(1, 1, "ISBT DEHRADUN")
sheet1.cell(2, 1, "SHASTRADHARA")
sheet1.cell(3, 1, "CLEMEN TOWN")
sheet1.cell(4, 1, "RAJPUR ROAD")
sheet1.cell(5, 1, "CLOCK TOWER")
sheet1.cell(1, 2, "ISBT DEHRADUN")
sheet1.cell(1, 3, "SHASTRADHARA")
sheet1.cell(1, 4, "CLEMEN TOWN")
sheet1.cell(1, 5, "RAJPUR ROAD")
sheet1.cell(1, 6, "CLOCK TOWER")

openpyxl_wb.save("openpyxl example.xlsx")
